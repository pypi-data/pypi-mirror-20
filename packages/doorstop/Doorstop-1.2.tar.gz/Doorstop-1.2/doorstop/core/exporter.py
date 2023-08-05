"""Functions to export documents and items."""

import os
import csv
import datetime
from collections import defaultdict

import yaml
import openpyxl

from doorstop import common
from doorstop.common import DoorstopError
from doorstop.core.types import iter_documents, iter_items
from doorstop import settings

LIST_SEP = '\n'  # string separating list values when joined in a string

XLSX_MAX_WIDTH = 65  # maximum width for a column
XLSX_FILTER_PADDING = 3.5  # column padding to account for filter button

log = common.logger(__name__)


def export(obj, path, ext=None, **kwargs):
    """Export an object to a given format.

    The function can be called in two ways:

    1. document or item-like object + output file path
    2. tree-like object + output directory path

    :param obj: (1) Item, list of Items, Document or (2) Tree
    :param path: (1) output file path or (2) output directory path
    :param ext: file extension to override output extension

    :raises: :class:`doorstop.common.DoorstopError` for unknown file formats

    :return: output location if files created, else None

    """
    # Determine the output format
    ext = ext or os.path.splitext(path)[-1] or '.csv'
    check(ext)

    # Export documents
    count = 0
    for obj2, path2 in iter_documents(obj, path, ext):
        count += 1

        # Export content to the specified path
        common.create_dirname(path2)
        log.info("exporting to {}...".format(path2))
        if ext in FORMAT_LINES:
            lines = export_lines(obj2, ext, **kwargs)
            common.write_lines(lines, path2)
        else:
            export_file(obj2, path2, ext, **kwargs)

    # Return the exported path
    if count:
        msg = "exported to {} file{}".format(count, 's' if count > 1 else '')
        log.info(msg)
        return path
    else:
        log.warning("nothing to export")
        return None


def export_lines(obj, ext='.yml', **kwargs):
    """Yield lines for an export in the specified format.

    :param obj: Item, list of Items, or Document to export
    :param ext: file extension to specify the output format

    :raises: :class:`doorstop.common.DoorstopError` for unknown file formats

    :return: lines generator

    """
    gen = check(ext, get_lines_gen=True)
    log.debug("yielding {} as lines of {}...".format(obj, ext))
    yield from gen(obj, **kwargs)


def export_file(obj, path, ext=None, **kwargs):
    """Create a file object for an export in the specified format.

    :param obj: Item, list of Items, or Document to export
    :param path: output file location with desired extension
    :param ext: file extension to override output path's extension

    :raises: :class:`doorstop.common.DoorstopError` for unknown file formats

    :return: path to created file

    """
    ext = ext or os.path.splitext(path)[-1]
    func = check(ext, get_file_func=True)
    log.debug("converting %s to file format %s...", obj, ext)
    try:
        return func(obj, path, **kwargs)
    except IOError:
        msg = "unable to write to: {}".format(path)
        raise common.DoorstopFileError(msg) from None


def _lines_yaml(obj, **_):
    """Yield lines for a YAML export.

    :param obj: Item, list of Items, or Document to export

    :return: iterator of lines of text

    """
    for item in iter_items(obj):

        data = {str(item.uid): item.data}
        text = yaml.dump(data, default_flow_style=False, allow_unicode=True)
        yield text


def _tabulate(obj, sep=LIST_SEP, auto=False):
    """Yield lines of header/data for tabular export.

    :param obj: Item, list of Items, or Document to export
    :param sep: string separating list values when joined in a string
    :param auto: include placeholders for new items on import

    :return: iterator of rows of data

    """
    yield_header = True

    for item in iter_items(obj):

        data = item.data

        # Yield header
        if yield_header:
            header = ['level', 'text', 'ref', 'links']
            for value in sorted(data.keys()):
                if value not in header:
                    header.append(value)
            yield ['uid'] + header
            yield_header = False

        # Yield row
        row = [item.uid]
        for key in header:
            value = data.get(key)
            if key == 'level':
                # some levels are floats for YAML presentation
                value = str(value)
            elif key == 'links':
                # separate identifiers with a delimiter
                value = sep.join(uid.string for uid in item.links)
            elif isinstance(value, str) and key not in ('reviewed',):
                # remove sentence boundaries and line wrapping
                value = item.get(key)
            elif value is None:
                value = ''
            row.append(value)
        yield row

    # Yield placeholders for new items
    if auto:
        for _ in range(settings.PLACEHOLDER_COUNT):
            yield [settings.PLACEHOLDER]


def _file_csv(obj, path, delimiter=',', auto=False):
    """Create a CSV file at the given path.

    :param obj: Item, list of Items, or Document to export
    :param path: location to export CSV file
    :param delimiter: character to delimit fields
    :param auto: include placeholders for new items on import

    :return: path of created file

    """
    with open(path, 'w', newline='', encoding='utf-8') as stream:
        writer = csv.writer(stream, delimiter=delimiter)
        for row in _tabulate(obj, auto=auto):
            writer.writerow(row)
    return path


def _file_tsv(obj, path, auto=False):
    """Create a TSV file at the given path.

    :param obj: Item, list of Items, or Document to export
    :param path: location to export TSV file
    :param auto: include placeholders for new items on import

    :return: path of created file

    """
    return _file_csv(obj, path, delimiter='\t', auto=auto)


def _file_xlsx(obj, path, auto=False):
    """Create an XLSX file at the given path.

    :param obj: Item, list of Items, or Document to export
    :param path: location to export XLSX file
    :param auto: include placeholders for new items on import

    :return: path of created file

    """
    workbook = _get_xlsx(obj, auto)
    workbook.save(path)

    return path


def _get_xlsx(obj, auto):
    """Create an XLSX workbook object.

    :param obj: Item, list of Items, or Document to export
    :param auto: include placeholders for new items on import

    :return: new workbook

    """
    col_widths = defaultdict(int)
    col = 'A'

    # Create a new workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Populate cells
    for row, data in enumerate(_tabulate(obj, auto=auto), start=1):
        for col_idx, value in enumerate(data, start=1):
            col = openpyxl.cell.get_column_letter(col_idx)
            cell = worksheet.cell('%s%s' % (col, row))

            # wrap text in every cell
            alignment = openpyxl.styles.Alignment(vertical='top',
                                                  horizontal='left',
                                                  wrap_text=True)
            style = cell.style.copy(alignment=alignment)
            # and bold header rows
            if row == 1:
                style = style.copy(font=openpyxl.styles.Font(bold=True))
            cell.style = style

            # convert incompatible Excel types:
            # http://pythonhosted.org/openpyxl/api.html#openpyxl.cell.Cell.value
            if not isinstance(value, (int, float, str, datetime.datetime)):
                value = str(value)
            cell.value = value

            # track cell width
            col_widths[col] = max(col_widths[col], _width(str(value)))

    # Add filter up to the last column
    worksheet.auto_filter.ref = "A1:%s1" % col

    # Set column width based on column contents
    for col in col_widths:
        if col_widths[col] > XLSX_MAX_WIDTH:
            width = XLSX_MAX_WIDTH
        else:
            width = col_widths[col] + XLSX_FILTER_PADDING
        worksheet.column_dimensions[col].width = width

    # Freeze top row
    worksheet.freeze_panes = worksheet.cell('A2')

    return workbook


def _width(text):
    """Get the maximum length in a multiline string."""
    if text:
        return max(len(line) for line in text.splitlines())
    else:
        return 0


# Mapping from file extension to lines generator
FORMAT_LINES = {'.yml': _lines_yaml}
# Mapping from file extension to file generator
FORMAT_FILE = {'.csv': _file_csv,
               '.tsv': _file_tsv,
               '.xlsx': _file_xlsx}
# Union of format dictionaries
FORMAT = dict(list(FORMAT_LINES.items()) + list(FORMAT_FILE.items()))


def check(ext, get_lines_gen=False, get_file_func=False):
    """Confirm an extension is supported for export.

    :param get_lines_func: return a lines generator if available
    :param get_file_func: return a file creator if available

    :raises: :class:`doorstop.common.DoorstopError` for unknown formats

    :return: function requested if available

    """
    exts = ', '.join(ext for ext in FORMAT)
    lines_exts = ', '.join(ext for ext in FORMAT_LINES)
    file_exts = ', '.join(ext for ext in FORMAT_FILE)
    fmt = "unknown {{}} format: {} (options: {{}})".format(ext or None)

    if get_lines_gen:
        try:
            gen = FORMAT_LINES[ext]
        except KeyError:
            exc = DoorstopError(fmt.format("lines export", lines_exts))
            raise exc from None
        else:
            log.debug("found lines generator for: {}".format(ext))
            return gen

    if get_file_func:
        try:
            func = FORMAT_FILE[ext]
        except KeyError:
            exc = DoorstopError(fmt.format("file export", file_exts))
            raise exc from None
        else:
            log.debug("found file creator for: {}".format(ext))
            return func

    if ext not in FORMAT:
        exc = DoorstopError(fmt.format("export", exts))
        raise exc
