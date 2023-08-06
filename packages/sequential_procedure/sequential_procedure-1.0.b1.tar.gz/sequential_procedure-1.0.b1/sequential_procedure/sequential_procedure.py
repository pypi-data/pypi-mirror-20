#import required libraries

from openpyxl import load_workbook
import numpy as np
import pandas as pd


class sample_size_estimator:

	def __init__(self, input_file, budget):
		#This is the input file to the webapp which contains 
	    #[1] Coeficients for each group
		#[2] Sampling cost for each group
		self.input_file = input_file 
		
		# total amount to be used to conduct experiment (experiment_budget)
		self.A = float(budget) 

	def getColumnValues(self, working_sheet, total_num_of_groups, column_index):
		columnVector = np.zeros(total_num_of_groups) # Coeficients for each group
		for i in range(0,total_num_of_groups):
			try:
				columnVector[i]=working_sheet.cell(row=i+2, column=column_index).value
				#print columnVector[i]
		 	except Exception as e: 
		 		print("Error : {} in function getColumnValues(a,b,c)".format(str(e)))
		return columnVector

	def performSequentialAnalysis(self):
		#read from excel file
		xlsx_file = pd.ExcelFile(self.input_file)
		#getting data stored in 'Sheet1;
		data = xlsx_file.parse('Sheet1')

		total_gene_ids = len(data['Gene'])

		#print("total_gene_ids : {}".format(total_gene_ids))

		mu = data['mu'].as_matrix()	# Mean for each group 
		sigma = data['sigma'].as_matrix() # Standard Deviation for each group
		C = data['Group-weightage'].as_matrix() # Coeficients for each group
		a = data['Gene-cost'].as_matrix() # Sampling cost for each group

		num_iterations=1
		current_iteration=1

		denominator_matrix = np.zeros((total_gene_ids, 1))
		test_result_matrix = np.zeros((total_gene_ids, num_iterations))

		#Theoritical calulation of sample value per group
 		optimum = np.zeros((total_gene_ids, 3))#optimum matrix containing both optimum sample size & respective deviation

 		#print C
 		#print sigma
 		#print sigma[0]

		for i in range(0,total_gene_ids):
			#print (self.A)*(np.absolute(C[i])*sigma[i])
			num = (self.A)*np.absolute(C[i])*sigma[i]
		 	for l in range(0,total_gene_ids):
		 		denominator_matrix[l]=np.absolute(C[l])*(sigma[l])*(np.sqrt(a[l]*a[i]))
		 	optimum[i] = num/np.sum(denominator_matrix)
		#print optimum

		while current_iteration <= num_iterations:
			print("Current iteration : {}".format(current_iteration))    	
			m=2#size of pilot data for each group
			counter = 0
			pilot_data = np.zeros((1, m))
			sd_of_each_group = np.zeros((total_gene_ids, 1)) #stores sd of each group when it reaches its optimum sample size
			#pilot_data generated successfully
			numerator=0
			denominator=0
			calculated_value = np.zeros((total_gene_ids, 1))#stores num/denom value for each group after each iteration
			state = np.zeros((total_gene_ids, 1))#this vector is used to determine whether that particular group achieved optimum value(i.e. 1) or not (i.e. "0")
			#generating pilot_data
			for i in range(0,total_gene_ids):
				temp = np.random.normal(mu[i],sigma[i], m)
				#print("Generated normal rand_num : {}".format(temp))    	
				#print("Pilot_data : {}".format(pilot_data))    	
				if (pilot_data[0]==0).all():
					pilot_data = temp
				else:
					pilot_data = np.vstack((pilot_data, temp))
			
			while counter < total_gene_ids:#this counter shows #groups that have achieved optimum value
		  		for i in range(0,total_gene_ids):
		  			if state[i]==0:
		  				numerator = (self.A)*abs(C[i])*(np.std(pilot_data[i]))
		   				for l in range(0,total_gene_ids):
							if state[l]==1:
								#print("IF:l : {} | {} ".format(l,pilot_data)) 
								denominator_matrix[l]=abs(C[l])*(sd_of_each_group[l])*np.sqrt(a[l]*a[i])
							else:#if(state[l]==0)
								#print("Else:l : {} | {}".format(l,pilot_data)) 
								denominator_matrix[l]=abs(C[l])*(np.std(pilot_data[l]))*np.sqrt(a[l]*a[i])
					calculated_value[i] = numerator/sum(denominator_matrix)
		    		###condition check
		  		for i in range(0,total_gene_ids):
					b = calculated_value[i]
					if (m>= b and state[i]==0):
						state[i]=1
						#optimum[i,1]=m
						#optimum[i,2]=sd(pilot_data[i,])
						test_result_matrix[i,(current_iteration-1)] = m
						sd_of_each_group[i]=np.std(pilot_data[i])
						counter = counter+1
									
								
				#steps to update m & pilot_data & restart process again for non-optimized groups
	  			m=m+1 #increamenting m
	  			pilot_data = np.hstack((pilot_data,np.zeros((total_gene_ids, 1))))
	  			for i in range(0,total_gene_ids):#updating the pilot data according to value of m
					if state[i]==0:
						pilot_data[i,(m-1)] = np.random.normal(mu[i], sigma[i], 1)
				
		   
		  	current_iteration = current_iteration + 1

  		for i in range(0,total_gene_ids): 
  			optimum[i,1] = np.mean(test_result_matrix[i])
   			optimum[i,2] = np.std(test_result_matrix[i])
		print(optimum)
		#print(test_result_matrix)
		return optimum

class sample_size_estimator_simulator:

	def __init__(self, input_file, budget, simulations):
		#This is the input file to the webapp which contains 
	    #[1] Coeficients for each group
		#[2] Sampling cost for each group
		self.input_file = input_file 
		
		# total amount to be used to conduct experiment (experiment_budget)
		self.A = float(budget) 

		# total number of simulations that need to be performed
		self.simulations=simulations

	def getColumnValues(self, working_sheet, total_num_of_groups, column_index):
		columnVector = np.zeros(total_num_of_groups) # Coeficients for each group
		for i in range(0,total_num_of_groups):
			try:
				columnVector[i]=working_sheet.cell(row=i+2, column=column_index).value
				#print columnVector[i]
		 	except Exception as e: 
		 		print("Error : {} in function getColumnValues(a,b,c)".format(str(e)))
		return columnVector

	def performSequentialAnalysis(self):
		#read from excel file
		xlsx_file = pd.ExcelFile(self.input_file)
		#getting data stored in 'Sheet1;
		data = xlsx_file.parse('Sheet1')

		total_gene_ids = len(data['Gene'])

		#print("total_gene_ids : {}".format(total_gene_ids))

		mu = data['mu'].as_matrix()	# Mean for each group 
		sigma = data['sigma'].as_matrix() # Standard Deviation for each group
		C = data['Group-weightage'].as_matrix() # Coeficients for each group
		a = data['Gene-cost'].as_matrix() # Sampling cost for each group

		num_iterations=self.simulations
		current_iteration=1

		denominator_matrix = np.zeros((total_gene_ids, 1))
		test_result_matrix = np.zeros((total_gene_ids, num_iterations))

		#Theoritical calulation of sample value per group
 		optimum = np.zeros((total_gene_ids, 3))#optimum matrix containing both optimum sample size & respective deviation

 		#print C
 		#print sigma
 		#print sigma[0]

		for i in range(0,total_gene_ids):
			#print (self.A)*(np.absolute(C[i])*sigma[i])
			num = (self.A)*np.absolute(C[i])*sigma[i]
		 	for l in range(0,total_gene_ids):
		 		denominator_matrix[l]=np.absolute(C[l])*(sigma[l])*(np.sqrt(a[l]*a[i]))
		 	optimum[i] = num/np.sum(denominator_matrix)
		#print optimum

		while current_iteration <= num_iterations:
			print("Current iteration : {}".format(current_iteration))    	
			m=2#size of pilot data for each group
			counter = 0
			pilot_data = np.zeros((1, m))
			sd_of_each_group = np.zeros((total_gene_ids, 1)) #stores sd of each group when it reaches its optimum sample size
			#pilot_data generated successfully
			numerator=0
			denominator=0
			calculated_value = np.zeros((total_gene_ids, 1))#stores num/denom value for each group after each iteration
			state = np.zeros((total_gene_ids, 1))#this vector is used to determine whether that particular group achieved optimum value(i.e. 1) or not (i.e. "0")
			#generating pilot_data
			for i in range(0,total_gene_ids):
				temp = np.random.normal(mu[i],sigma[i], m)
				#print("Generated normal rand_num : {}".format(temp))    	
				#print("Pilot_data : {}".format(pilot_data))    	
				if (pilot_data[0]==0).all():
					pilot_data = temp
				else:
					pilot_data = np.vstack((pilot_data, temp))
			
			while counter < total_gene_ids:#this counter shows #groups that have achieved optimum value
		  		for i in range(0,total_gene_ids):
		  			if state[i]==0:
		  				numerator = (self.A)*abs(C[i])*(np.std(pilot_data[i]))
		   				for l in range(0,total_gene_ids):
							if state[l]==1:
								#print("IF:l : {} | {} ".format(l,pilot_data)) 
								denominator_matrix[l]=abs(C[l])*(sd_of_each_group[l])*np.sqrt(a[l]*a[i])
							else:#if(state[l]==0)
								#print("Else:l : {} | {}".format(l,pilot_data)) 
								denominator_matrix[l]=abs(C[l])*(np.std(pilot_data[l]))*np.sqrt(a[l]*a[i])
					calculated_value[i] = numerator/sum(denominator_matrix)
		    		###condition check
		  		for i in range(0,total_gene_ids):
					b = calculated_value[i]
					if (m>= b and state[i]==0):
						state[i]=1
						#optimum[i,1]=m
						#optimum[i,2]=sd(pilot_data[i,])
						test_result_matrix[i,(current_iteration-1)] = m
						sd_of_each_group[i]=np.std(pilot_data[i])
						counter = counter+1
									
								
				#steps to update m & pilot_data & restart process again for non-optimized groups
	  			m=m+1 #increamenting m
	  			pilot_data = np.hstack((pilot_data,np.zeros((total_gene_ids, 1))))
	  			for i in range(0,total_gene_ids):#updating the pilot data according to value of m
					if state[i]==0:
						pilot_data[i,(m-1)] = np.random.normal(mu[i], sigma[i], 1)
				
		   
		  	current_iteration = current_iteration + 1

  		for i in range(0,total_gene_ids): 
  			optimum[i,1] = np.mean(test_result_matrix[i])
   			optimum[i,2] = np.std(test_result_matrix[i])
		print(optimum)
		#print(test_result_matrix)
		return optimum

		 	
