import os
import ntpath
import xlrd
import openpyxl
from openpyxl.utils import coordinate_from_string, column_index_from_string
from openpyxl.utils.exceptions import CellCoordinatesException


class ExcelManager:
    """
    Wrapper that opens and operates on .xls, .xlsx or .xlsm excel files. By
    default we take in a string representing the excel file path (extension
    included), and depending on the file type use xlrd or openpyxl to operate
    on it.

    The dev facing api is identical for either - internally we use xlrd or
    openpyxl methods depending on the file type.

    For rows & columns we use 1 based indexing to stay with the more modern
    openpyxl (and most users are more familiar with it if they are coming from
    an office environment, not a programming one). Be aware of which file type
    you are using if you retrieve the sheet object - it could be using zero OR
    one based indexing.

    Public Variables:
    file_path: full file path with extension of the file we are operating on
    workbook: openpyxl/xlrd workbook object for this file
    sheet: currently in use openpxl/xlrd sheet object for this work book
    read_count: number of sheet reads this object has done
    write_count: number of sheet writes this object has done

    Public Methods:
    select_sheet - choose which sheet to use (by index or name)
    cell - retrieve an openpyxl/xlrd cell object by row/column or index
    read - retrieve the value from the current sheet at row/column or index
    write - write a value to the current sheet at row/column or index
    save - save the workbook at the initial file path, or a new file path
    if one is specified
    info - return basic information/status of the object
    to_array - return a 2D numpy array representation of the current sheet
    find_index - return the first index of the match or None if it does not exist
    find_indexes - return a list of tuples containing the indexes of all matches
    """

    write_count = 0
    read_count = 0
    sheet_array = None

    def __init__(self, file_path, sheet_name=None, sheet_index=None):
        self.file_path = file_path
        self.__check_file_extension(file_path)
        self.__check_file_exists(file_path)
        if file_path.endswith('.xls'):
            self.__is_xls = True
            self.__init_xls(sheet_name, sheet_index)
        else:
            self.__is_xls = False
            self.__init_excel(sheet_name, sheet_index)

    def change_sheet(self, *args):
        """
        Change the current active sheet object
        :param name: sheet name
        :param index: sheet index (1 index)
        :return: None
        """
        if isinstance(args[0], str):
            name = args[0]
            index = None
        elif isinstance(args[0], int):
            name = None
            index = args[0]
        else:
            raise ValueError('Specify either the sheet name or sheet index to change sheets')

        if self.__is_xls:
            self.__select_xls_sheet(name, index - 1 if index else None)
        else:
            self.__select_excel_sheet(name, index - 1 if index else None)

    def row(self, row_index):
        """
        Return the row at the specified index
        :row_index row_index: 1 based index
        :return: list of values
        """
        self.sheet_array = self.array()
        return self.sheet_array[row_index - 1]

    def column(self, column_index):
        """
        return the column at the specified index
        :param column_index: string or (1 based) int index
        :return: list of values
        """
        if isinstance(column_index, int):
            column = column_index - 1
        else:
            column = column_index_from_string(column_index.upper()) - 1
        self.sheet_array = self.array()
        return [row[column] for row in self.sheet_array]

    def cell(self, *args):
        """
        Return the cell at the specified location
        :param args: tuple with either a 1 based representation for row/column
        or string based index
        :return: xlrd/openpyxl cell object
        """
        row, column = self.__parse_row_column_from_args(*args)
        if self.__is_xls:
            return self.__get_xls_cell(row - 1, column - 1)  # xlrd is a 1 based index
        else:
            return self.__get_excel_cell(row, column)

    def read(self, *args):
        """
        Read the value from the target cell
        :param args: tuple with either a 1 based representation for row/column
        or string based index
        :return: string
        """
        self.read_count += 1
        value = self.cell(*args).value
        return value if value else ''

    def write(self, *args, value=None):
        """
        Input the value at the specified target
        :param args: tuple with either a 1 based representation for row/column
        or string based index
        :param value:
        :return:
        """
        if self.__is_xls:
            raise TypeError('Writing to a cell is not supported for .xls files')
        self.cell(*args).value = value
        self.write_count += 1

    def save(self, *args):
        """
        Save the current sheet either at the original file_path (if none
        specified) or at the file_path parameter
        :param file_path: new file path to save file
        :return: None
        """
        if len(args) == 1:
            self.__check_file_extension(args[0])
            file_path = args[0]
        else:
            file_path = self.file_path
        if self.__is_xls:
            raise TypeError('Saving is not supported for .xls files')
        else:
            self.workbook.save(file_path)

    def info(self, string=False):
        """
        return basic information about this ExcelWrapper instance
        :return: string
        """
        sheet_name = self.sheet.name if self.__is_xls else self.sheet.title
        if string:
            return 'File: {}\nSheet: {}\nReads: {}\nWrites: {}' \
                .format(self.file_path, sheet_name, self.read_count, self.write_count)
        else:
            return {
                'file': self.file_path,
                'sheet': sheet_name,
                'reads': self.read_count,
                'writes': self.write_count
            }

    def array(self):
        """
        Return a 2D list representing the spreadsheet
        :return: list(list())
        """
        if self.__is_xls:
            self.sheet_array = self.__xls_to_array()
            return self.sheet_array
        else:
            self.sheet_array = self.__excel_to_array()
            return self.sheet_array

    def search(self, value, match=1, case_insensitive=True, contains=False, many=False):
        """
        Given a value find the 1 based index where that value is located
        on the sheet or None if it does not exist. If 'many' is set true then
        an empty list is returned if no matches are found
        :param value: the value we are searching for
        :param match: if multiple results are found we return only one - this
        parameter determines which index of the list we return with a 1 based index
        :param case_insensitive: whether or not the search should be case insensitive
        :param contains: whether or not the search should use 'in' or equality to
        check if the spreadsheet value is a match
        :param many: whether or not to return a singular value or a list of values
        :return:
        """
        indexes = self.__find_indexes(value, case_insensitive=case_insensitive, contains=contains)
        if many:
            return indexes
        try:
            match = indexes[match - 1]
            return match[0], match[1]
        except IndexError:
            return None, None

    def __find_indexes(self, value, case_insensitive, contains):
        """
        Iterate over the 2D list representation of the sheet and determine
        if the input value exists based on search parameters
        :param value: value we are looking for
        :param case_insensitive: whether or not search is case_insensitive
        :param contains: use 'in' to find matches
        :return:
        """
        self.sheet_array = self.array()
        indexes = []
        for i, row in enumerate(self.sheet_array):
            for j, column in enumerate(row):
                input_val, column_val = self.__check_case_sensitive(case_insensitive, value, column)
                if contains and input_val in column_val:
                    indexes.append((i + 1, j + 1))
                elif input_val == column_val:
                    indexes.append((i + 1, j + 1))
        return indexes

    @staticmethod
    def __check_case_sensitive(case_insensitive, value, column):
        column_val = column.lower() if case_insensitive else column
        input_val = value.lower() if case_insensitive else value
        return input_val, column_val


    def __parse_row_column_from_args(self, *args):
        """
        convert a generic arguments tuple into a 1-based ow/column index. This
        is to support both numeric (1, 1) and string (A1) representation of
        cells with the same API.
        :param args: args tuple
        :return: int, int tuple
        """
        if len(args) == 1 and isinstance(args[0], str):
            row, column = self.__parse_row_column_from_index(args[0])
        elif len(args) == 2 and isinstance(args[0], int) and isinstance(args[1], int):
            row = args[0]
            column = args[1]
        else:
            raise ValueError('Specify either row and column numbers (1, 1) OR a cell index ("A1")')
        return row, column

    @staticmethod
    def __parse_row_column_from_index(cell_index):
        """
        Given a string based excel index return the int based row, column
        representation
        :param cell_index: string based excel input
        :return: row, column ints
        """
        try:
            xy = coordinate_from_string(cell_index.upper())
            row = xy[1]
            column = column_index_from_string(xy[0])
            return row, column
        except CellCoordinatesException:
            raise ValueError('The index must be a valid Excel index (A1, E17, etc.)')

    def __init_xls(self, sheet_name, sheet_index):
        """
        initialize a .xls file with xlrd
        """
        self.workbook = xlrd.open_workbook(self.file_path)
        self.__select_xls_sheet(sheet_name, sheet_index)

    def __select_xls_sheet(self, sheet_name, sheet_index):
        """
        change the currently active xlrd sheet
        """
        if sheet_name:
            self.sheet = self.workbook.sheet_by_name(sheet_name)
        elif sheet_index:
            self.sheet = self.workbook.sheet_by_index(sheet_index)
        else:
            self.sheet = self.workbook.sheet_by_index(0)

    def __get_xls_cell(self, row, column):
        """
        retrieve the xlrd cell object at the specified row/column
        :param row: 1-based row index
        :param column: 1-based column index
        :return: cell object
        """
        return self.sheet.cell(row, column)

    def __xls_to_array(self):
        """
        convert an xlrd sheet to a 2D list of values.
        :return:
        """
        sheet_array = []
        for row in range(1, self.__get_max_rows() + 1):
            row_array = []
            for column in range(1, self.__get_max_columns() + 1):
                value = self.read(row, column)
                row_array.append(value)
            sheet_array.append(row_array)
        return sheet_array

    def __init_excel(self, sheet_name, sheet_index):
        """
        initialize a .xlsx file with openpyxl
        """
        self.workbook = openpyxl.load_workbook(self.file_path)
        self.__select_excel_sheet(sheet_name, sheet_index)

    def __select_excel_sheet(self, sheet_name, sheet_index):
        """
        change the currently active openpyxl sheet
        """
        if sheet_name:
            self.sheet = self.workbook[sheet_name]
        elif sheet_index:
            sheet_names = self.workbook.sheetnames
            self.sheet = self.workbook[sheet_names[sheet_index]]
        else:
            sheet_names = self.workbook.sheetnames
            self.sheet = self.workbook[(sheet_names[0])]

    def __get_excel_cell(self, row, column):
        """
        retrieve the openpyxl cell object at the specified row/column
        :param row: 1-based row index
        :param column: 1-based column index
        :return: cell object
        """
        return self.sheet.cell(row=row, column=column)

    def __excel_to_array(self):
        """
        convert an openpyxl sheet to a 2D list of values.
        :return:
        """
        sheet_array = []
        for row in range(1, self.sheet.max_row + 1):
            row_array = []
            for column in range(1, self.sheet.max_column + 1):
                value = self.read(row, column)
                row_array.append(value)
            sheet_array.append(row_array)
        return sheet_array

    def __get_max_rows(self):
        """
        return the number of rows in the current xlrd sheet object
        :return: int
        """
        if self.__is_xls:
            return self.sheet.nrows
        return self.sheet.max_rows

    def __get_max_columns(self):
        """
        return the number of columns in the current xlrd sheet object
        :return: int
        """
        if self.__is_xls:
            return self.sheet.ncols
        return self.sheet.max_column

    @staticmethod
    def __check_file_extension(file_path):
        extensions = ['.xls', '.xlsx', '.xlsm']
        if not any(file_path.endswith(extension) for extension in extensions):
            raise ValueError("""
            No extension found on file path - make sure you include the FULL file path with the extension. Valid
            extensions include: {}
            """.format(', '.join(extensions)))

    @staticmethod
    def __check_file_exists(file_path):
        """
        Check to see if the input file exists - if not, raise an error
        that lists other excel files in the same directory
        :param file_path: full file path to excel file
        :return:
        """
        if not os.path.exists(file_path):
            file_name = ntpath.basename(file_path)
            file_directory = file_path.replace(file_name, '')
            valid_files = [file for file in os.listdir(file_directory) if 'xls' in file]
            base_error = 'The file {} was not found. Maybe you were looking for one of these?\n\n'.format(file_name)
            raise FileNotFoundError(base_error + '\n'.join(valid_files[:10]))


